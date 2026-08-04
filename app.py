import os
import csv
import json
import smtplib
import random
import time
import re
import hashlib
import hmac
import html
import threading
import base64
import urllib.parse
import urllib.request
import urllib.error
import uuid
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.utils import secure_filename
from flask import Flask, render_template, send_from_directory, request, jsonify, Response, abort, session, redirect
from flask_cors import CORS
from openpyxl import Workbook, load_workbook

otp_store = {}
registration_lock = threading.Lock()


def _verification_secret():
    return (os.environ.get('VERIFICATION_SECRET') or app.secret_key).encode('utf-8')


def create_email_verification_token(email):
    """Create a short-lived, server-verifiable email verification token."""
    payload = json.dumps({
        'email': email.strip().lower(),
        'expires': int(time.time()) + 1800,
    }, separators=(',', ':')).encode('utf-8')
    encoded = base64.urlsafe_b64encode(payload).decode().rstrip('=')
    signature = hmac.new(_verification_secret(), encoded.encode('ascii'), hashlib.sha256).hexdigest()
    return f'{encoded}.{signature}'


def verify_email_verification_token(email, token):
    try:
        encoded, signature = token.split('.', 1)
        expected = hmac.new(_verification_secret(), encoded.encode('ascii'), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            return False
        padding = '=' * (-len(encoded) % 4)
        payload = json.loads(base64.urlsafe_b64decode((encoded + padding).encode('ascii')))
        return payload.get('email') == email.strip().lower() and int(payload.get('expires', 0)) >= int(time.time())
    except (ValueError, TypeError, KeyError, json.JSONDecodeError, UnicodeError):
        return False


app = Flask(__name__, template_folder='public/templates', static_folder='public/assets')
CORS(app, supports_credentials=True) # Allow cross-origin requests from Firebase
app.secret_key = 'super-secret-key-surya'
app.config.update(
    SESSION_COOKIE_SAMESITE='None',
    SESSION_COOKIE_SECURE=True
)

# Basic Auth credentials
ADMIN_USER = 'admin'
ADMIN_PASS = 'changeme'

UPLOAD_FOLDER = os.path.join('public', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Helper to check authentication
def is_authenticated():
    return session.get('admin_logged_in') is True

def _check_auth(auth):
    return auth and auth.username == ADMIN_USER and auth.password == ADMIN_PASS

# Context processor to make broadcasts available globally
@app.context_processor
def inject_broadcasts():
    return dict(broadcasts=load_broadcasts())

def load_csv_rows(filename):
    if not os.path.exists(filename):
        return []
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            return list(reader)
    except Exception as e:
        print(f"Error loading {filename}: {e}")
        return []

def load_broadcasts():
    broadcasts = []
    if os.path.exists('broadcasts.csv'):
        try:
            with open('broadcasts.csv', 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    broadcasts.append(row)
        except Exception as e:
            print(f"Error loading broadcasts.csv: {e}")
    return list(reversed(broadcasts))

# Helper to load environmental variables
def load_dotenv():
    if os.path.exists('.env'):
        with open('.env', 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                parts = line.split('=', 1)
                if len(parts) == 2:
                    os.environ[parts[0].strip()] = parts[1].strip()

load_dotenv()

# Helper to send emails
def send_allocation_email(candidate_email, candidate_name, reg_id, exam, center):
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASSWORD')
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    try:
        port = int(os.environ.get('SMTP_PORT', '587'))
    except:
        port = 587

    email_body = f"""
    <html>
      <head>
        <style>
          body {{ font-family: 'Segoe UI', Arial, sans-serif; background-color: #f1f5f9; color: #1e293b; margin: 0; padding: 20px; }}
          .card {{ background-color: #ffffff; max-width: 600px; margin: 0 auto; border-radius: 12px; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1); overflow: hidden; border: 1px solid #e2e8f0; }}
          .header {{ background: linear-gradient(135deg, #1e3a8a, #3b82f6); color: white; padding: 24px; text-align: center; }}
          .content {{ padding: 24px; line-height: 1.6; }}
          .detail-box {{ background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px; margin: 16px 0; }}
          .footer {{ text-align: center; font-size: 0.8rem; color: #64748b; padding: 16px; border-top: 1px solid #e2e8f0; }}
          .badge {{ background-color: #3b82f6; color: white; padding: 4px 8px; border-radius: 4px; font-weight: bold; }}
        </style>
      </head>
      <body>
        <div class="card">
          <div class="header">
            <h2>SURYA EDUCATION BOARD</h2>
            <p>Exam Center Allocation Notice</p>
          </div>
          <div class="content">
            <p>Dear <strong>{candidate_name}</strong>,</p>
            <p>Your exam center has been successfully allocated. Please find the details of your registration and allocated center below:</p>
            <div class="detail-box">
              <p><strong>Registration ID:</strong> <code style="color: #1e3a8a; font-weight: bold;">{reg_id}</code></p>
              <p><strong>Exam Category:</strong> {exam}</p>
              <p><strong>Allocated Center:</strong> <span class="badge">{center}</span></p>
            </div>
            <p>Please print your admission slip from the candidate portal to carry with you to the exam center.</p>
          </div>
          <div class="footer">
            <p>© 2026 SURYA Education Board. All rights reserved.</p>
          </div>
        </div>
      </body>
    </html>
    """

    if not smtp_user or not smtp_pass:
        print("[SMTP LOG] No SMTP credentials configured. Printing email content to console:", flush=True)
        print(f"[SMTP LOG] To: {candidate_email}", flush=True)
        print(f"[SMTP LOG] Subject: SURYA Exam Center Allocated - {reg_id}", flush=True)
        print("[SMTP LOG] Body:\n", email_body, flush=True)
        return

    def send_thread():
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f"SURYA Exam Center Allocated - {reg_id}"
            msg['From'] = f"SURYA Education Board <{smtp_user}>"
            msg['To'] = candidate_email
            msg.attach(MIMEText(email_body, 'html'))

            server = smtplib.SMTP(smtp_host, port)
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, candidate_email, msg.as_string())
            server.quit()
            print(f"[SMTP LOG] Email successfully sent to {candidate_email} for reg {reg_id}", flush=True)
        except Exception as e:
            print(f"[SMTP LOG] Error sending email: {e}", flush=True)

    import threading
    threading.Thread(target=send_thread).start()

def send_otp_email(candidate_email, otp):
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASSWORD')
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    try:
        port = int(os.environ.get('SMTP_PORT', '587'))
    except:
        port = 587

    email_body = f"""
    <html>
      <head>
        <style>
          body {{ font-family: 'Segoe UI', Arial, sans-serif; background-color: #f1f5f9; color: #1e293b; margin: 0; padding: 20px; }}
          .card {{ background-color: #ffffff; max-width: 600px; margin: 0 auto; border-radius: 12px; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1); overflow: hidden; border: 1px solid #e2e8f0; }}
          .header {{ background: linear-gradient(135deg, #1e3a8a, #3b82f6); color: white; padding: 24px; text-align: center; }}
          .content {{ padding: 24px; line-height: 1.6; }}
          .detail-box {{ background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px; margin: 16px 0; text-align: center; font-size: 24px; letter-spacing: 4px; }}
          .footer {{ text-align: center; font-size: 0.8rem; color: #64748b; padding: 16px; border-top: 1px solid #e2e8f0; }}
        </style>
      </head>
      <body>
        <div class="card">
          <div class="header">
            <h2>SURYA EDUCATION BOARD</h2>
            <p>Email Verification</p>
          </div>
          <div class="content">
            <p>Dear Candidate,</p>
            <p>Your One Time Password (OTP) for registration is:</p>
            <div class="detail-box">
              <strong>{otp}</strong>
            </div>
            <p>This OTP is valid for 10 minutes. Do not share it with anyone.</p>
          </div>
          <div class="footer">
            <p>© 2026 SURYA Education Board. All rights reserved.</p>
          </div>
        </div>
      </body>
    </html>
    """

    if not smtp_user or not smtp_pass:
        print("[SMTP LOG] No SMTP credentials configured. Printing OTP to console:", flush=True)
        print(f"[SMTP LOG] To: {candidate_email}, OTP: {otp}", flush=True)
        return

    def send_thread():
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f"SURYA Registration OTP: {otp}"
            msg['From'] = f"SURYA Education Board <{smtp_user}>"
            msg['To'] = candidate_email
            msg.attach(MIMEText(email_body, 'html'))

            if port == 465:
                server = smtplib.SMTP_SSL(smtp_host, port)
                server.login(smtp_user, smtp_pass)
            else:
                server = smtplib.SMTP(smtp_host, port)
                server.starttls()
                server.login(smtp_user, smtp_pass)
            
            server.sendmail(smtp_user, candidate_email, msg.as_string())
            server.quit()
            print(f"[SMTP LOG] OTP email sent to {candidate_email}", flush=True)
        except Exception as e:
            print(f"[SMTP LOG] Error sending OTP email: {e}", flush=True)

    import threading
    threading.Thread(target=send_thread).start()

def send_gmail_otp(candidate_email, otp):
    """Send an OTP through Gmail SMTP and report whether Gmail accepted it."""
    smtp_user = (os.environ.get('GMAIL_USER') or os.environ.get('SMTP_USER') or '').strip()
    smtp_pass = (os.environ.get('GMAIL_APP_PASSWORD') or os.environ.get('SMTP_PASSWORD') or '').replace(' ', '').strip()
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com').strip()
    try:
        port = int(os.environ.get('SMTP_PORT', '587'))
    except ValueError:
        port = 587

    message = MIMEMultipart('alternative')
    message['Subject'] = 'SURYA Registration Email Verification OTP'
    message['From'] = f'SURYA Education Board <{smtp_user}>'
    message['To'] = candidate_email
    message.attach(MIMEText(
        f'<p>Dear Candidate,</p><p>Your SURYA registration OTP is:</p>'
        f'<h2 style="letter-spacing:4px">{otp}</h2>'
        '<p>This OTP expires in 10 minutes. Do not share it with anyone.</p>',
        'html'
    ))

    gmail_client_id = os.environ.get('GMAIL_CLIENT_ID', '').strip()
    gmail_client_secret = os.environ.get('GMAIL_CLIENT_SECRET', '').strip()
    gmail_refresh_token = os.environ.get('GMAIL_REFRESH_TOKEN', '').strip()
    if gmail_client_id and gmail_client_secret and gmail_refresh_token:
        try:
            token_payload = urllib.parse.urlencode({
                'client_id': gmail_client_id,
                'client_secret': gmail_client_secret,
                'refresh_token': gmail_refresh_token,
                'grant_type': 'refresh_token',
            }).encode()
            token_request = urllib.request.Request(
                'https://oauth2.googleapis.com/token', data=token_payload,
                headers={'Content-Type': 'application/x-www-form-urlencoded'}, method='POST')
            with urllib.request.urlopen(token_request, timeout=10) as response:
                access_token = json.loads(response.read().decode())['access_token']
            raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode().rstrip('=')
            gmail_request = urllib.request.Request(
                'https://gmail.googleapis.com/gmail/v1/users/me/messages/send',
                data=json.dumps({'raw': raw_message}).encode(),
                headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'},
                method='POST')
            with urllib.request.urlopen(gmail_request, timeout=10):
                pass
            print(f'[GMAIL API] OTP accepted for {candidate_email}', flush=True)
            return True, ''
        except (urllib.error.HTTPError, urllib.error.URLError, KeyError, json.JSONDecodeError) as exc:
            print(f'[GMAIL API] OTP failed for {candidate_email}: {exc}', flush=True)
            return False, 'Gmail API rejected the request. Check the OAuth client and refresh token in Render.'

    if not smtp_user or not smtp_pass:
        return False, 'Configure Gmail OAuth credentials for HTTPS delivery or SMTP credentials on the server.'

    ports = [port] if port == 465 else [port, 465]
    last_error = None
    for attempt_port in ports:
        server = None
        try:
            if attempt_port == 465:
                server = smtplib.SMTP_SSL(smtp_host, attempt_port, timeout=10)
            else:
                server = smtplib.SMTP(smtp_host, attempt_port, timeout=10)
                server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, [candidate_email], message.as_string())
            server.quit()
            print(f'[SMTP LOG] Gmail OTP accepted for {candidate_email}', flush=True)
            return True, ''
        except smtplib.SMTPAuthenticationError as exc:
            last_error = exc
            if server:
                server.quit()
            return False, 'Gmail rejected the login. Generate a new Gmail App Password and update the Render variable.'
        except (smtplib.SMTPConnectError, TimeoutError, OSError) as exc:
            last_error = exc
            if server:
                server.close()
            continue
        except Exception as exc:
            last_error = exc
            if server:
                server.close()
            break

    print(f'[SMTP LOG] Gmail OTP failed for {candidate_email}: {last_error}', flush=True)
    return False, 'Render could not connect to Gmail SMTP on ports 587 or 465. Check SMTP connectivity or use a transactional email provider.'

def send_registration_email(candidate_email, candidate_name, application_number, registration_data, password, pdf_url):
    """Send the candidate a confirmation containing their login details."""
    smtp_user = (os.environ.get('GMAIL_USER') or os.environ.get('SMTP_USER') or '').strip()
    smtp_pass = (os.environ.get('GMAIL_APP_PASSWORD') or os.environ.get('SMTP_PASSWORD') or '').replace(' ', '').strip()
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    try:
        port = int(os.environ.get('SMTP_PORT', '587'))
    except ValueError:
        port = 587

    safe_name = html.escape(candidate_name)
    detail_rows = []
    for key, raw_value in registration_data.items():
        if key == 'password_hash' or key.startswith('file_') or raw_value in ('', None):
            continue
        label = html.escape(str(key).replace('-', ' ').replace('_', ' ').title())
        detail = html.escape(str(raw_value))
        detail_rows.append(f'<tr><td style="padding:6px;font-weight:bold">{label}</td><td style="padding:6px">{detail}</td></tr>')
    detail_table = ''.join(detail_rows)
    public_base_url = os.environ.get('PUBLIC_API_BASE_URL', 'https://surya-s2f5.onrender.com').rstrip('/')
    safe_pdf_url = html.escape(public_base_url + pdf_url)
    logo_url = html.escape(public_base_url + '/assets/favicon.svg')
    logo = f'<img src="{logo_url}" width="42" height="42" alt="SURYA logo" style="display:inline-block;border-radius:12px">'
    email_body = f"""
    <html><body style="margin:0;background:#f1f5f9;font-family:Arial,sans-serif;color:#1e293b;padding:24px">
      <div style="max-width:680px;margin:auto;background:#fff;border:1px solid #dbe3ef;border-radius:16px;overflow:hidden">
        <div style="background:linear-gradient(135deg,#172554,#2563eb);color:#fff;padding:24px;text-align:center">
          {logo}<h1 style="margin:10px 0 4px;font-size:22px">SURYA EDUCATION BOARD</h1>
          <p style="margin:0;color:#dbeafe">Registration Confirmation</p>
        </div>
        <div style="padding:26px">
          <p>Dear <strong>{safe_name}</strong>,</p>
          <p>Your registration has been completed successfully. Keep this email for your records.</p>
          <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;padding:16px;margin:20px 0">
            <p style="margin:0 0 8px"><strong>Application Number:</strong> {html.escape(application_number)}</p>
            <p style="margin:0"><strong>Password:</strong> {html.escape(password)}</p>
          </div>
          <h3 style="color:#1e3a8a;margin-bottom:8px">Submitted Details</h3>
          <table width="100%" border="0" cellpadding="0" cellspacing="0" style="border-collapse:collapse;border:1px solid #cbd5e1">
            {detail_table}
          </table>
          <p style="text-align:center;margin:24px 0">
            <a href="{safe_pdf_url}" style="display:inline-block;background:#2563eb;color:#fff;text-decoration:none;padding:12px 20px;border-radius:8px;font-weight:bold">Download Application PDF</a>
          </p>
          <p style="font-size:12px;color:#64748b">Keep your application number and password private.</p>
        </div>
        <div style="background:#f8fafc;border-top:1px solid #e2e8f0;padding:14px;text-align:center;color:#64748b;font-size:12px">© 2026 SURYA Education Board</div>
      </div>
    </body></html>
    """

    gmail_client_id = os.environ.get('GMAIL_CLIENT_ID', '').strip()
    gmail_client_secret = os.environ.get('GMAIL_CLIENT_SECRET', '').strip()
    gmail_refresh_token = os.environ.get('GMAIL_REFRESH_TOKEN', '').strip()
    if gmail_client_id and gmail_client_secret and gmail_refresh_token:
        try:
            token_payload = urllib.parse.urlencode({'client_id': gmail_client_id, 'client_secret': gmail_client_secret, 'refresh_token': gmail_refresh_token, 'grant_type': 'refresh_token'}).encode()
            token_request = urllib.request.Request('https://oauth2.googleapis.com/token', data=token_payload, headers={'Content-Type': 'application/x-www-form-urlencoded'}, method='POST')
            with urllib.request.urlopen(token_request, timeout=10) as response:
                access_token = json.loads(response.read().decode())['access_token']
            email_message = MIMEMultipart('alternative')
            email_message['Subject'] = f"SURYA Registration Confirmation - {application_number}"
            email_message['From'] = f"SURYA Education Board <{smtp_user}>"
            email_message['To'] = candidate_email
            email_message.attach(MIMEText(email_body, 'html'))
            raw_message = base64.urlsafe_b64encode(email_message.as_bytes()).decode().rstrip('=')
            gmail_request = urllib.request.Request('https://gmail.googleapis.com/gmail/v1/users/me/messages/send', data=json.dumps({'raw': raw_message}).encode(), headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}, method='POST')
            with urllib.request.urlopen(gmail_request, timeout=10):
                pass
            print(f"[GMAIL API] Registration email sent to {candidate_email}", flush=True)
            return True, ''
        except Exception as exc:
            print(f"[GMAIL API] Registration email failed for {candidate_email}: {exc}", flush=True)

    if not smtp_user or not smtp_pass:
        print(f"[EMAIL LOG] Registration email not sent: Gmail API or SMTP credentials are missing (recipient: {candidate_email})", flush=True)
        return False, 'Email service credentials are not configured'

    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"SURYA Registration Confirmation - {application_number}"
        msg['From'] = f"SURYA Education Board <{smtp_user}>"
        msg['To'] = candidate_email
        msg.attach(MIMEText(email_body, 'html'))
        if port == 465:
            server = smtplib.SMTP_SSL(smtp_host, port, timeout=15)
            server.login(smtp_user, smtp_pass)
        else:
            server = smtplib.SMTP(smtp_host, port, timeout=15)
            server.starttls()
            server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, [candidate_email], msg.as_string())
        server.quit()
        print(f"[SMTP LOG] Registration email sent to {candidate_email}", flush=True)
        return True, ''
    except Exception as exc:
        print(f"[SMTP LOG] Registration email failed for {candidate_email}: {exc}", flush=True)
        return False, 'The confirmation email could not be sent. Please contact support.'

def generate_application_pdf(application_number, registration_data, password, upload_dir):
    """Create the candidate's downloadable application summary PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    pdf_path = os.path.join(upload_dir, 'application-form.pdf')
    styles = getSampleStyleSheet()
    body = styles['BodyText']
    body.fontSize = 9
    body.leading = 12
    title = styles['Title']
    title.textColor = colors.HexColor('#1e3a8a')
    rows = [
        [Paragraph('<b>Application Number</b>', body), Paragraph(html.escape(application_number), body)],
        [Paragraph('<b>Password</b>', body), Paragraph(html.escape(password), body)],
    ]
    for key, raw_value in registration_data.items():
        if key == 'password_hash' or key.startswith('file_') or raw_value in ('', None):
            continue
        label = str(key).replace('-', ' ').replace('_', ' ').title()
        rows.append([Paragraph(f'<b>{html.escape(label)}</b>', body), Paragraph(html.escape(str(raw_value)), body)])

    document = SimpleDocTemplate(
        pdf_path, pagesize=A4, rightMargin=16 * mm, leftMargin=16 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
    )
    story = [Paragraph('SURYA EDUCATION BOARD', title), Paragraph('Candidate Application Form', styles['Heading2']), Spacer(1, 8)]
    table = Table(rows, colWidths=[55 * mm, 115 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#eaf2ff')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#94a3b8')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 7), ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph('Please keep this application form for your records.', body))
    document.build(story)
    return f'/uploads/{application_number}/application-form.pdf'

def registration_duplicate(data):
    """Return the duplicate field label, if this person has already registered."""
    candidate = {
        'reg-email': value_normalized(data.get('reg-email')),
        'reg-mobile': value_normalized(data.get('reg-mobile')),
        'pd-id': value_normalized(data.get('pd-id')),
    }
    records = []
    if os.path.exists('seoas_registrations.jsonl'):
        try:
            with open('seoas_registrations.jsonl', 'r', encoding='utf-8') as file:
                records.extend(json.loads(line) for line in file if line.strip())
        except (OSError, json.JSONDecodeError):
            pass
    if os.path.exists('seoas_registrations.xlsx'):
        try:
            workbook = load_workbook('seoas_registrations.xlsx', read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = list(next(rows, ()))
            records.extend(dict(zip(headers, row)) for row in rows)
            workbook.close()
        except (OSError, ValueError):
            pass
    labels = {'reg-email': 'email address', 'reg-mobile': 'mobile number', 'pd-id': 'government ID'}
    for record in records:
        for key, current in candidate.items():
            if current and value_normalized(record.get(key)) == current:
                return labels[key]
    return None

def value_normalized(raw):
    return re.sub(r'\s+', '', str(raw or '')).strip().lower()

# Explicitly serve static assets and uploads
@app.route('/assets/<path:path>')
def send_assets(path):
    return send_from_directory('public/assets', path)

@app.route('/uploads/<path:path>')
def send_uploads(path):
    return send_from_directory('public/uploads', path)

# API Root Health Check
@app.route('/')
def home():
    return jsonify({"status": "Surya Backend API is running. Please access the frontend via Firebase."})

# API routes
@app.route('/api/send-otp', methods=['POST'])
def send_otp():
    contact = request.form.get('contact', '').strip()
    contact_type = request.form.get('type', '').strip()
    
    if not contact or not contact_type:
        return jsonify({"status": "error", "message": "Contact and type are required"}), 400
    if contact_type not in {'Email', 'Mobile'}:
        return jsonify({"status": "error", "message": "Unsupported verification type"}), 400
    if contact_type == 'Email' and not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', contact):
        return jsonify({"status": "error", "message": "Enter a valid email address"}), 400
        
    otp = str(random.randint(100000, 999999))
    otp_store[contact] = {
        "otp": otp,
        "expires": time.time() + 600, # 10 minutes validity
        "type": contact_type,
    }
    
    if contact_type == 'Email':
        sent, error_message = send_gmail_otp(contact, otp)
        if not sent:
            otp_store.pop(contact, None)
            return jsonify({"status": "error", "message": error_message}), 502
    elif contact_type == 'Mobile':
        # TODO: Implement SMS provider (e.g. Twilio) here
        print(f"[SMS LOG] To: {contact}, OTP: {otp}", flush=True)
        
    return jsonify({"status": "ok", "message": f"OTP sent to {contact}"})

@app.route('/api/verify-otp', methods=['POST'])
def verify_otp():
    contact = request.form.get('contact', '').strip()
    otp = request.form.get('otp', '').strip()
    
    if not contact or not otp:
        return jsonify({"status": "error", "message": "Contact and OTP are required"}), 400
        
    record = otp_store.get(contact)
    if not record:
        return jsonify({"status": "error", "message": "No OTP requested for this contact"}), 400
        
    if time.time() > record["expires"]:
        del otp_store[contact]
        return jsonify({"status": "error", "message": "OTP has expired"}), 400
        
    if record["otp"] == otp:
        del otp_store[contact]
        response = {"status": "ok", "message": "Verification successful"}
        if record.get('type') == 'Email':
            response['verification_token'] = create_email_verification_token(contact)
        return jsonify(response)
    else:
        return jsonify({"status": "error", "message": "Invalid OTP"}), 400

@app.route('/api/seoas-register', methods=['POST'])
def seoas_register():
    """Validate a registration, save uploaded files, and append the row to Excel."""
    try:
        data_str = request.form.get('data', '{}')
        data = json.loads(data_str)
    except (TypeError, json.JSONDecodeError):
        return jsonify({"status": "error", "message": "Invalid JSON data"}), 400

    if not isinstance(data, dict):
        return jsonify({"status": "error", "message": "Registration data must be an object"}), 400

    def value(key):
        return str(data.get(key, '')).strip()

    email_token = request.form.get('email_verification_token', '').strip()
    if not email_token or not verify_email_verification_token(value('reg-email'), email_token):
        return jsonify({"status": "error", "message": "Please verify the registered email address before submitting"}), 400

    required = {
        'reg-name': 'Candidate name', 'reg-mobile': 'Mobile number',
        'reg-email': 'Email address', 'reg-pass': 'Password',
        'reg-pass2': 'Confirm password', 'pd-gender': 'Gender',
        'pd-dob': 'Date of birth', 'pd-id': 'Government ID',
        'pd-cat': 'Category', 'pd-pwd': 'PwD status',
        'pd-father': "Father's name", 'pd-mother': "Mother's name",
        'pd-income': 'Annual income', 'pd-address': 'Permanent address',
        'pd-state': 'State', 'pd-district': 'District', 'pd-pin': 'PIN code',
        'exam-medium': 'Exam medium',
        'preference-1': 'Centre preference 1', 'preference-2': 'Centre preference 2'
    }
    missing = [label for key, label in required.items() if not value(key)]
    if missing:
        return jsonify({"status": "error", "message": "Please complete: " + ', '.join(missing)}), 400

    password = value('reg-pass')
    if not re.fullmatch(r'(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z\d]).{8,}', password):
        return jsonify({"status": "error", "message": "Password must be at least 8 characters and include uppercase, lowercase, number, and special character"}), 400
    if password != value('reg-pass2'):
        return jsonify({"status": "error", "message": "Password and confirm password do not match"}), 400
    if not re.fullmatch(r'[6-9]\d{9}', value('reg-mobile')):
        return jsonify({"status": "error", "message": "Enter a valid 10-digit Indian mobile number"}), 400
    if not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', value('reg-email')):
        return jsonify({"status": "error", "message": "Enter a valid email address"}), 400
    if not re.fullmatch(r'\d{12}', value('pd-id')):
        return jsonify({"status": "error", "message": "Government ID must contain 12 digits"}), 400
    if not re.fullmatch(r'\d{6}', value('pd-pin')):
        return jsonify({"status": "error", "message": "PIN code must contain 6 digits"}), 400
    academic_values = {key: value(key) for key in data.keys() if str(key).startswith('academic') and value(key)}
    if not academic_values:
        return jsonify({"status": "error", "message": "Please complete the academic details"}), 400
    if value('academic-year'):
        try:
            year = int(value('academic-year'))
            if year < 1950 or year > datetime.now().year:
                raise ValueError
        except ValueError:
            return jsonify({"status": "error", "message": "Enter a valid passing year"}), 400
    if value('academic-marks'):
        try:
            marks = float(value('academic-marks'))
            if marks < 0 or marks > 100:
                raise ValueError
        except ValueError:
            return jsonify({"status": "error", "message": "Enter marks between 0 and 100"}), 400

    allowed_centers = {
        'Bengaluru', 'Mysuru', 'Mangaluru', 'Hubballi-Dharwad', 'Kalaburagi',
        'Chennai', 'Coimbatore', 'Madurai', 'Tiruchirappalli', 'Salem',
        'Mumbai', 'Pune', 'Nagpur', 'Nashik', 'Aurangabad (Chhatrapati Sambhajinagar)',
        'Hyderabad', 'Warangal', 'Karimnagar', 'Nizamabad', 'Khammam',
        'Visakhapatnam', 'Vijayawada', 'Tirupati', 'Guntur', 'Kurnool',
        'Thiruvananthapuram', 'Kochi', 'Kozhikode', 'Thrissur', 'Kannur'
    }
    if value('preference-1') not in allowed_centers or value('preference-2') not in allowed_centers:
        return jsonify({"status": "error", "message": "Select exam cities from the approved test centers list"}), 400

    with registration_lock:
        duplicate = registration_duplicate(data)
    if duplicate:
        return jsonify({"status": "error", "message": f"A registration already exists for this {duplicate}. Only one registration is allowed per person."}), 409

    app_no = f"SEOAS-{datetime.now().year}-{uuid.uuid4().hex[:12].upper()}"
    data['application_number'] = app_no
    data['timestamp'] = datetime.now().isoformat()
    data['password_hash'] = hashlib.sha256(password.encode('utf-8')).hexdigest()
    data.pop('reg-pass', None)
    data.pop('reg-pass2', None)
    
    # Handle files
    upload_dir = os.path.join('public', 'uploads', app_no)
    os.makedirs(upload_dir, exist_ok=True)
    
    for key in request.files:
        file = request.files[key]
        if file.filename:
            filename = secure_filename(file.filename)
            filepath = os.path.join(upload_dir, filename)
            file.save(filepath)
            data[f'file_{key}'] = f'/uploads/{app_no}/{filename}'
            
    # Save an audit copy and append a human-readable Excel row.
    try:
        with open('seoas_registrations.jsonl', 'a', encoding='utf-8') as f:
            f.write(json.dumps(data) + '\n')
        excel_path = 'seoas_registrations.xlsx'
        headers = list(data.keys())
        if os.path.exists(excel_path):
            workbook = load_workbook(excel_path)
            sheet = workbook.active
            existing_headers = [cell.value for cell in sheet[1]]
            for header in headers:
                if header not in existing_headers:
                    existing_headers.append(header)
                    sheet.cell(row=1, column=len(existing_headers), value=header)
            headers = existing_headers
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Registrations'
            for column, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=column, value=header)
            sheet.freeze_panes = 'A2'
            sheet.auto_filter.ref = f'A1:{chr(64 + min(len(headers), 26))}1'
        row = sheet.max_row + 1
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=row, column=column, value=data.get(header, ''))
        workbook.save(excel_path)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
        
    try:
        pdf_url = generate_application_pdf(app_no, data, password, upload_dir)
    except Exception as exc:
        print(f'[PDF] Failed to create application PDF for {app_no}: {exc}', flush=True)
        return jsonify({"status": "error", "message": "Application saved, but the PDF could not be generated. Please contact support."}), 500

    email_sent, email_error = send_registration_email(value('reg-email'), value('reg-name'), app_no, data, password, pdf_url)
    public_details = {key: value for key, value in data.items() if key != 'password_hash'}
    message = "Application submitted successfully. A confirmation email has been sent." if email_sent else "Application submitted successfully, but the confirmation email could not be sent. Please download the PDF below and contact support."
    return jsonify({"status": "ok", "application_number": app_no, "pdf_url": pdf_url, "details": public_details, "password": password, "email_sent": email_sent, "email_error": email_error, "message": message})

@app.route('/register', methods=['POST'])
def handle_register():
    return jsonify(status='error', message='Registrations for 2027 exams is not yet announced/open.'), 403

@app.route('/contact', methods=['POST'])
def handle_contact():
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    subject = request.form.get('subject', '').strip()
    message = request.form.get('message', '').strip()

    if not name or not email or not subject or not message:
        return jsonify(status='error', message='Missing required fields'), 400

    file_exists = os.path.exists('contacts.csv')
    try:
        with open('contacts.csv', 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(['name', 'email', 'subject', 'message'])
            writer.writerow([name, email, subject, message])
        return jsonify(status='ok')
    except Exception as e:
        return jsonify(status='error', message=str(e)), 500

@app.route('/api/verify-candidate', methods=['POST'])
def verify_candidate():
    application_number = request.form.get('application_number', '').strip().upper()
    password = request.form.get('password', '').strip()

    if not application_number or not password:
        return jsonify(status='error', message='Application Number and password are required'), 400

    # Dummy verification
    candidate = {
        'application_number': application_number,
        'name': 'Dummy Candidate',
        'email': 'dummy@example.com'
    }
    return jsonify(status='ok', candidate=candidate)

# Admin Portal
@app.route('/admin')
def admin():
    auth = request.authorization
    if auth and _check_auth(auth):
        session['admin_logged_in'] = True
    
    if not session.get('admin_logged_in'):
        return Response(
            'Could not verify your access level for that URL.\n'
            'You have to login with proper credentials', 401,
            {'WWW-Authenticate': 'Basic realm="Admin Required"'}
        )

    registration_rows = load_csv_rows('registrations.csv')
    contact_rows = load_csv_rows('contacts.csv')
    
    return render_template(
        'admin.html',
        registration_rows=registration_rows,
        contact_rows=contact_rows,
        broadcasts=load_broadcasts(),
        registration_closed=True
    )

@app.route('/admin/download-registrations')
def download_registrations():
    auth = request.authorization
    if not (session.get('admin_logged_in') or (auth and _check_auth(auth))):
        return Response(
            'Authentication required', 401,
            {'WWW-Authenticate': 'Basic realm="Admin Required"'}
        )
    
    if not os.path.exists('registrations.csv'):
        return 'No registrations found', 404
        
    try:
        with open('registrations.csv', 'r', encoding='utf-8') as f:
            content = f.read()
        return Response(
            content,
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment;filename=registrations.csv'}
        )
    except Exception as e:
        return str(e), 500

@app.route('/admin/allocate-center', methods=['POST'])
def allocate_center():
    auth = request.authorization
    if not (session.get('admin_logged_in') or (auth and _check_auth(auth))):
        return jsonify(status='error', message='Unauthorized'), 401

    reg_id = request.form.get('registration_id')
    center = request.form.get('allocated_center')

    if not reg_id or not center:
        return jsonify(status='error', message='Missing parameter'), 400

    if not os.path.exists('registrations.csv'):
        return jsonify(status='error', message='No registrations file'), 404

    updated = False
    rows = []
    headers = []
    candidate_email = ''
    candidate_name = ''
    exam_category = ''

    try:
        with open('registrations.csv', 'r', encoding='utf-8', newline='') as f:
            reader = csv.reader(f)
            rows = list(reader)
            if rows:
                headers = rows[0]

        reg_id_idx = headers.index('registration_id') if 'registration_id' in headers else 0
        center_idx = headers.index('allocated_center') if 'allocated_center' in headers else 10
        email_idx = headers.index('email') if 'email' in headers else 2
        name_idx = headers.index('name') if 'name' in headers else 1
        exam_idx = headers.index('exam_category') if 'exam_category' in headers else 6

        for i in range(1, len(rows)):
            if rows[i][reg_id_idx] == reg_id:
                rows[i][center_idx] = center
                candidate_email = rows[i][email_idx]
                candidate_name = rows[i][name_idx]
                exam_category = rows[i][exam_idx]
                updated = True
                break

        if updated:
            with open('registrations.csv', 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            if center != 'Pending' and candidate_email:
                send_allocation_email(candidate_email, candidate_name, reg_id, exam_category, center)
            return jsonify(status='ok')
        else:
            return jsonify(status='error', message='Registration ID not found'), 404
    except Exception as e:
        return jsonify(status='error', message=str(e)), 500

@app.route('/admin/broadcast', methods=['POST'])
def handle_broadcast():
    auth = request.authorization
    if not (session.get('admin_logged_in') or (auth and _check_auth(auth))):
        return jsonify(status='error', message='Unauthorized'), 401

    title = request.form.get('title', '').strip()
    content = request.form.get('content', '').strip()

    if not title or not content:
        return jsonify(status='error', message='Missing required fields'), 400

    doc_file = request.files.get('document')
    doc_filename = ''
    if doc_file and doc_file.filename != '':
        try:
            filename = secure_filename(doc_file.filename)
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            doc_file.save(os.path.join(UPLOAD_FOLDER, filename))
            doc_filename = filename
        except Exception as e:
            return jsonify(status='error', message=f"Failed to save document: {str(e)}"), 500

    file_exists = os.path.exists('broadcasts.csv')
    next_id = 1
    
    if file_exists:
        try:
            with open('broadcasts.csv', 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        next_id = max(next_id, int(row.get('id', 0)) + 1)
                    except:
                        pass
        except Exception as e:
            print(f"Error reading broadcasts.csv for ID: {e}")

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    try:
        with open('broadcasts.csv', 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(['id', 'date', 'title', 'content', 'document'])
            writer.writerow([next_id, now_str, title, content, doc_filename])
        
        with open('public/broadcasts.csv', 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if not os.path.exists('public/broadcasts.csv'):
                writer.writerow(['id', 'date', 'title', 'content', 'document'])
            writer.writerow([next_id, now_str, title, content, doc_filename])
            
        return jsonify(status='ok')
    except Exception as e:
        return jsonify(status='error', message=str(e)), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
